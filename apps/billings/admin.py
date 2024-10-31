from django.contrib import admin
from datetime import date, timedelta
from django.utils.translation import gettext as _
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from apps.billings.models import Billing, BillingProduct

# Register your models here.
class ProductTabularInline(admin.TabularInline):
    model = BillingProduct
    extra = 0

def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="billings.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Billings"

    columns = ['ID', 'Итоговая цена', 'Цена доставки', 'Адрес', 'Номер телефона', 'Оплата', 'Код оплаты', 'Статус', 'Создан']
    ws.append(columns)

    column_widths = [len(column) for column in columns]

    for obj in queryset:
        status_display = "Оплачено" if obj.status else "Не оплачено"
        
        row = [
            obj.id, obj.total_price, obj.delivery_price, obj.address, obj.phone,
            obj.payment_code, status_display, obj.created.strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row)
        for i, cell in enumerate(row):
            column_widths[i] = max(column_widths[i], len(str(cell)))

    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width + 2  

    wb.save(response)
    return response

export_to_excel.short_description = "Export to Excel"

@admin.register(Billing)
class BillingAdmin(admin.ModelAdmin):
    list_display = ('id', 'delivery_date_time', 'total_price', 'first_name', 'last_name', 'payment_code', 'billing_receipt_type', 'billing_status', 'created')
    search_fields = ('id', 'delivery_date_time', 'total_price', 'first_name', 'last_name', 'payment_code', 'billing_receipt_type', 'billing_status')
    ordering = ('-created', )
    inlines = [ProductTabularInline]
    list_filter = ('billing_receipt_type', )
    actions = [export_to_excel]

@admin.register(BillingProduct) 
class BillingProductAdmin(admin.ModelAdmin):
    list_display = ('id', 'billing', 'product', 'quantity', 'price', 'status')